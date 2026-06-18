"""
output_writer.py — Módulo de escrita de resultados em múltiplos formatos.

Suporta: CSV, TSV, JSON e XLSX.
Consolida resultados de múltiplos arquivos em um único relatório.
"""

import csv
import io
import json
import logging

logger = logging.getLogger(__name__)


def parse_tsv_data(tsv_text: str) -> tuple[list[str], list[list[str]]]:
    """
    Parseia dados TSV da IA em cabeçalho + linhas.

    Returns:
        Tupla (cabeçalho, lista_de_linhas)
    """
    lines = [l for l in tsv_text.strip().split("\n") if l.strip()]

    if not lines:
        return [], []

    reader = csv.reader(lines, delimiter="\t", quoting=csv.QUOTE_ALL)
    rows = list(reader)

    if len(rows) == 0:
        return [], []

    header = rows[0]
    data = rows[1:] if len(rows) > 1 else []

    return header, data


def write_csv(header: list[str], data: list[list[str]], separator: str = ",") -> str:
    """Gera conteúdo CSV/TSV como string."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=separator, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(header)
    for row in data:
        writer.writerow(row)
    return output.getvalue()


def write_json(header: list[str], data: list[list[str]]) -> str:
    """Gera conteúdo JSON estruturado como string."""
    records = []
    for row in data:
        record = {}
        for i, col in enumerate(header):
            record[col] = row[i] if i < len(row) else ""
        records.append(record)

    return json.dumps(records, ensure_ascii=False, indent=2)


def write_xlsx(header: list[str], data: list[list[str]], file_path: str):
    """Escreve dados em um arquivo Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "Biblioteca 'openpyxl' não instalada. "
            "Execute: pip install openpyxl"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Análise"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Escrever cabeçalho
    for col, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escrever dados
    data_alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, min(len(str(cell.value)), 50))
        adjusted_width = max_length + 4
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Congelar a linha do cabeçalho
    ws.freeze_panes = "A2"

    wb.save(file_path)
    logger.info(f"XLSX salvo em: {file_path}")


def save_result(
    tsv_data: str,
    file_path: str,
    output_format: str = "csv",
) -> str:
    """
    Salva os dados analisados no formato escolhido.

    Args:
        tsv_data: Dados TSV brutos da IA
        file_path: Caminho para salvar o arquivo
        output_format: "csv", "tsv", "json" ou "xlsx"

    Returns:
        Caminho do arquivo salvo
    """
    header, data = parse_tsv_data(tsv_data)

    if not header:
        raise ValueError("Dados vazios — nada para salvar")

    if output_format == "csv":
        content = write_csv(header, data, separator=",")
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            f.write(content)

    elif output_format == "tsv":
        content = write_csv(header, data, separator="\t")
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            f.write(content)

    elif output_format == "json":
        content = write_json(header, data)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

    elif output_format == "xlsx":
        write_xlsx(header, data, file_path)

    else:
        raise ValueError(f"Formato '{output_format}' não suportado")

    logger.info(f"Resultado salvo como {output_format.upper()}: {file_path}")
    return file_path


def merge_batch_results(results: list[tuple[str, str, str | None]]) -> str:
    """
    Consolida resultados de múltiplos arquivos em um único TSV.

    Mantém o cabeçalho do primeiro resultado e adiciona uma coluna
    "Arquivo de Origem" para rastreabilidade.

    Args:
        results: Lista de (nome_arquivo, dados_tsv, erro_ou_none)

    Returns:
        TSV consolidado com todos os resultados
    """
    all_headers = None
    all_data = []

    for filename, tsv_data, error in results:
        if error or not tsv_data:
            continue

        header, data = parse_tsv_data(tsv_data)

        if not header:
            continue

        if all_headers is None:
            all_headers = ["Arquivo de Origem"] + header

        for row in data:
            all_data.append([filename] + row)

    if not all_headers or not all_data:
        return ""

    return write_csv(all_headers, all_data, separator="\t")
